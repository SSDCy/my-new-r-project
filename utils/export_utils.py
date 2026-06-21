from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
import pandas as pd
import numpy as np

def add_df_to_sheet(wb, df, sheet_name, index=False):
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=index, header=True):
        ws.append(r)
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    return ws

def get_group_colors(groups):
    """从分组名称生成颜色映射，与原 R 代码一致"""
    palette = ['#FFB6C1','#90EE90','#87CEEB','#DDA0DD','#FFD700',
               '#FFA07A','#98FB98','#B0C4DE','#FFB347','#C9A0DC']
    if len(groups) == 0:
        return {}
    if len(groups) > len(palette):
        # 扩展颜色
        import matplotlib.pyplot as plt
        cmap = plt.cm.tab20
        colors = [cmap(i % 20) for i in range(len(groups))]
        return {g: f'#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}' for g, (r,g,b,_) in zip(groups, colors)}
    return {g: palette[i] for i, g in enumerate(groups)}

def create_excel_report(raw_data, clean_data, norm_data, de_results, comparisons, groups, sample_info):
    """
    生成完整的 Excel 报告，对齐原 R 代码逻辑：
    - 多个 sheet：Raw Data, Clean Data, Normalized, 每个比较单独 sheet, DE Summary
    - 分组颜色应用到强度列
    - 差异表根据 regulation 着色
    - 冻结首行
    """
    print("[DEBUG] 开始生成 Excel 报告")
    wb = Workbook()
    wb.remove(wb.active)

    # 数据 sheets
    add_df_to_sheet(wb, raw_data, "Raw Data")
    add_df_to_sheet(wb, clean_data, "Clean Data")
    add_df_to_sheet(wb, norm_data, "Normalized Data")

    # 分组颜色映射（基于 groups 字典的值）
    # groups 格式: {group_name: [col1, col2, ...]}
    group_color_map = {}
    if groups:
        group_names = list(groups.keys())
        colors = get_group_colors(group_names)
        for gn, cols in groups.items():
            for col in cols:
                # 在 Norm 数据中，可能有 Norm_ 前缀的列
                norm_col = 'Norm_' + col
                group_color_map[col] = colors[gn]
                group_color_map[norm_col] = colors[gn]

    # 为 Normalized 和 Clean 数据添加颜色（可选）
    for sheet_name in ["Normalized Data", "Clean Data"]:
        ws = wb[sheet_name]
        # 找到第一行的列名
        headers = [cell.value for cell in ws[1]]
        for idx, header in enumerate(headers, start=1):
            if header in group_color_map:
                fill = PatternFill(start_color=group_color_map[header].lstrip('#'), end_color=group_color_map[header].lstrip('#'), fill_type="solid")
                for row in ws.iter_rows(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = fill
        # 冻结首行
        ws.freeze_panes = 'A2'

    # 差异分析结果 sheets
    color_map_reg = {
        'Up': 'FF0000',
        'Down': '0000FF',
        'Increase': 'C00000',
        'Decrease': '0945A5',
        'NS': '7f7e83'
    }

    for comp_name, res_df in de_results.items():
        print(f"[DEBUG] 写入比较: {comp_name}")
        # 添加注释列（来自 norm_data 的 Protein IDs, Majority protein IDs, Unique peptides 等）
        annotation_cols = []
        for col in ['Protein IDs', 'Majority protein IDs', 'Master protein IDs', 'Unique peptides']:
            if col in norm_data.columns:
                annotation_cols.append(col)
        # 合并
        if annotation_cols and 'Master protein IDs' in annotation_cols:
            # 以 Master protein IDs 为键合并
            norm_sub = norm_data[annotation_cols].drop_duplicates(subset='Master protein IDs')
            merged = res_df.merge(norm_sub, on='Master protein IDs', how='left')
        else:
            merged = res_df

        ws = add_df_to_sheet(wb, merged, comp_name[:31])

        # 为 regulation 列着色
        reg_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'regulation':
                reg_col_idx = idx
                break
        if reg_col_idx:
            for row in ws.iter_rows(min_col=reg_col_idx, max_col=reg_col_idx, min_row=2, max_row=ws.max_row):
                cell = row[0]
                reg = str(cell.value)
                if reg in color_map_reg:
                    cell.font = Font(color=color_map_reg[reg])
        # 冻结首行
        ws.freeze_panes = 'A2'

    # DE Summary  sheet
    summary = []
    for comp_name, res_df in de_results.items():
        up = (res_df['regulation'] == 'Up').sum()
        down = (res_df['regulation'] == 'Down').sum()
        inc = (res_df['regulation'] == 'Increase').sum()
        dec = (res_df['regulation'] == 'Decrease').sum()
        summary.append({
            'Comparison': comp_name,
            'Up': up,
            'Down': down,
            'Increase': inc,
            'Decrease': dec,
            'Total': up + down + inc + dec
        })
    if summary:
        add_df_to_sheet(wb, pd.DataFrame(summary), "DE Summary")

    print("[DEBUG] Excel 报告生成完毕")
    return wb
